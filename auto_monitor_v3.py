#!/usr/bin/env python3
"""
Autonomous Trading Monitor v3.0
- No daily trade limit (for testing)
- Quality filters increase with trade count
- Adaptive entry criteria
"""

import json
import time
import random
from datetime import datetime
from openpyxl import load_workbook

class AdaptiveTrader:
    def __init__(self):
        self.capital = 100000
        self.max_risk_per_trade = 1500
        self.max_daily_loss = 1500
        self.max_positions = 2
        self.current_positions = []
        self.daily_pnl = 0
        self.trades_today = 0
        self.consecutive_losses = 0
        
        # Adaptive quality thresholds
        self.base_entry_min = 1.5  # Base minimum gain %
        self.base_entry_max = 3.0  # Base maximum gain %
        
        self.watchlist = [
            {"symbol": "TATASTEEL", "sector": "Metals", "base_price": 155.0, "volatility": 0.010, "trend_bias": 0.005},
            {"symbol": "NTPC", "sector": "Power", "base_price": 355.0, "volatility": 0.007, "trend_bias": 0.003},
            {"symbol": "AXISBANK", "sector": "Banking", "base_price": 1145.0, "volatility": 0.009, "trend_bias": 0.004},
            {"symbol": "BEL", "sector": "Defense", "base_price": 295.0, "volatility": 0.011, "trend_bias": 0.006},
            {"symbol": "SBIN", "sector": "Banking", "base_price": 820.0, "volatility": 0.008, "trend_bias": 0.003},
            {"symbol": "COALINDIA", "sector": "Metals", "base_price": 425.0, "volatility": 0.006, "trend_bias": 0.002},
            {"symbol": "ONGC", "sector": "Energy", "base_price": 245.0, "volatility": 0.007, "trend_bias": 0.003},
            {"symbol": "POWERGRID", "sector": "Power", "base_price": 315.0, "volatility": 0.006, "trend_bias": 0.002},
        ]
        
        # Re-entry tracking
        self.stopped_stocks = {}  # {symbol: timestamp}
        
    def get_current_time(self):
        return datetime.now().strftime("%H:%M:%S")
    
    def get_quality_threshold(self):
        """Increase quality requirements as trade count grows"""
        if self.trades_today <= 3:
            # First 3 trades: Standard criteria
            return {
                'entry_min': 1.5,
                'entry_max': 3.0,
                'probability': 0.70,
                'min_sector_strength': 0.0
            }
        elif self.trades_today <= 6:
            # Trades 4-6: Tighter criteria
            return {
                'entry_min': 1.5,
                'entry_max': 2.5,  # Avoid extended stocks
                'probability': 0.80,  # Higher conviction needed
                'min_sector_strength': 0.5
            }
        elif self.trades_today <= 10:
            # Trades 7-10: Very selective
            return {
                'entry_min': 1.5,
                'entry_max': 2.0,  # Only fresh moves
                'probability': 0.85,
                'min_sector_strength': 1.0
            }
        else:
            # 10+ trades: Extremely selective
            return {
                'entry_min': 1.5,
                'entry_max': 1.8,  # Very fresh only
                'probability': 0.90,
                'min_sector_strength': 1.5
            }
    
    def check_reentry_allowed(self, symbol):
        """Check if re-entry is allowed based on cooldown"""
        if symbol not in self.stopped_stocks:
            return True
        
        stopped_time = self.stopped_stocks[symbol]
        cooldown_mins = 30  # 30 minute cooldown
        
        # For simulation, we'll use a simple counter
        # In real system, would check actual time elapsed
        return False  # For now, prevent re-entry same day
    
    def load_positions(self):
        try:
            wb = load_workbook('trades.xlsx')
            sheet = wb['Trade Log']
            positions = []
            
            for row in range(2, sheet.max_row + 1):
                outcome = sheet[f'O{row}'].value
                if outcome == "OPEN":
                    pos = {
                        'symbol': sheet[f'C{row}'].value,
                        'entry': sheet[f'F{row}'].value,
                        'sl': sheet[f'G{row}'].value,
                        'target': sheet[f'H{row}'].value,
                        'qty': sheet[f'I{row}'].value,
                        'row': row
                    }
                    positions.append(pos)
            
            self.current_positions = positions
            self.trades_today = sheet.max_row - 1
            
            # Count consecutive losses
            self.consecutive_losses = 0
            for row in range(sheet.max_row, 1, -1):
                outcome = sheet[f'O{row}'].value
                if outcome == "LOSS":
                    self.consecutive_losses += 1
                elif outcome in ["WIN", "OPEN"]:
                    break
            
            return positions
        except Exception as e:
            print(f"Error loading positions: {e}")
            return []
    
    def simulate_price_movement(self, symbol, base_price=None):
        stock_data = next((s for s in self.watchlist if s['symbol'] == symbol), None)
        
        if stock_data:
            volatility = stock_data['volatility']
            trend_bias = stock_data['trend_bias']
            price = stock_data['base_price'] if base_price is None else base_price
        elif symbol == "HINDALCO":
            volatility = 0.008
            trend_bias = 0.004
            price = 648.0 if base_price is None else base_price
        else:
            volatility = 0.006
            trend_bias = 0.002
            price = base_price if base_price else 100.0
        
        random_move = random.gauss(trend_bias, volatility)
        new_price = price * (1 + random_move)
        
        return round(new_price, 2)
    
    def check_position_status(self, position, current_price):
        entry = position['entry']
        sl = position['sl']
        target = position['target']
        
        if current_price <= sl:
            return "SL_HIT", current_price
        elif current_price >= target:
            return "TARGET_HIT", current_price
        elif current_price > entry:
            profit_pct = ((current_price - entry) / entry) * 100
            if profit_pct >= 1.0:
                return "TRAILING", current_price
        
        return "OPEN", current_price
    
    def update_position(self, position, status, exit_price):
        try:
            wb = load_workbook('trades.xlsx')
            sheet = wb['Trade Log']
            row = position['row']
            
            sheet[f'L{row}'] = exit_price
            qty = position['qty']
            entry = position['entry']
            pnl = (exit_price - entry) * qty
            pnl_pct = ((exit_price - entry) / entry) * 100
            
            sheet[f'M{row}'] = round(pnl, 2)
            sheet[f'N{row}'] = f"{pnl_pct:.2f}%"
            
            if status == "SL_HIT":
                sheet[f'O{row}'] = "LOSS"
                self.stopped_stocks[position['symbol']] = datetime.now()
            elif status == "TARGET_HIT":
                sheet[f'O{row}'] = "WIN"
            
            wb.save('trades.xlsx')
            
            self.daily_pnl += pnl
            return pnl
            
        except Exception as e:
            print(f"Error updating position: {e}")
            return 0
    
    def calculate_position_size(self, entry_price, sl_price):
        sl_distance = abs(entry_price - sl_price)
        if sl_distance == 0:
            return 0
        
        # After 2 consecutive losses, reduce risk
        risk_multiplier = 0.67 if self.consecutive_losses >= 2 else 1.0
        adjusted_risk = self.max_risk_per_trade * risk_multiplier
        
        max_shares = int(adjusted_risk / sl_distance)
        position_value = max_shares * entry_price
        
        if position_value > 40000:
            max_shares = int(40000 / entry_price)
        
        return max_shares
    
    def execute_new_trade(self, stock):
        symbol = stock['symbol']
        current_price = self.simulate_price_movement(symbol)
        
        entry = current_price
        sl = round(current_price * 0.985, 2)
        target = round(current_price * 1.027, 2)
        
        qty = self.calculate_position_size(entry, sl)
        
        if qty == 0:
            print(f"  ⚠️ Cannot calculate position size for {symbol}")
            return False
        
        capital_used = qty * entry
        risk_amount = qty * (entry - sl)
        rr_ratio = (target - entry) / (entry - sl)
        
        try:
            wb = load_workbook('trades.xlsx')
            sheet = wb['Trade Log']
            row = sheet.max_row + 1
            
            sheet[f'A{row}'] = "16-Jan-2026"
            sheet[f'B{row}'] = self.get_current_time()
            sheet[f'C{row}'] = symbol
            sheet[f'D{row}'] = "LONG"
            sheet[f'E{row}'] = f"{stock['sector']} Strength / Adaptive Entry"
            sheet[f'F{row}'] = entry
            sheet[f'G{row}'] = sl
            sheet[f'H{row}'] = target
            sheet[f'I{row}'] = qty
            sheet[f'J{row}'] = round(capital_used, 2)
            sheet[f'K{row}'] = f"1:{rr_ratio:.1f}"
            sheet[f'O{row}'] = "OPEN"
            
            wb.save('trades.xlsx')
            
            risk_note = " (REDUCED RISK)" if self.consecutive_losses >= 2 else ""
            
            print(f"\n🎯 TRADE #{self.trades_today + 1} EXECUTED: {symbol} LONG{risk_note}")
            print(f"  Entry: ₹{entry} | SL: ₹{sl} | Target: ₹{target}")
            print(f"  Qty: {qty} shares | Capital: ₹{capital_used:,.0f}")
            print(f"  Risk: ₹{risk_amount:.0f} | R:R: 1:{rr_ratio:.1f}")
            
            self.trades_today += 1
            return True
            
        except Exception as e:
            print(f"  ❌ Error executing trade: {e}")
            return False
    
    def scan_and_execute(self):
        if len(self.current_positions) >= self.max_positions:
            return None
        
        quality = self.get_quality_threshold()
        
        print(f"\n🔍 SCANNING (Quality Mode: Trade #{self.trades_today + 1})")
        print(f"   Entry Range: {quality['entry_min']:.1f}%-{quality['entry_max']:.1f}%")
        print(f"   Conviction: {quality['probability']*100:.0f}%+")
        if self.consecutive_losses >= 2:
            print(f"   ⚠️ DEFENSIVE MODE: {self.consecutive_losses} consecutive losses - reduced sizing")
        
        for stock in self.watchlist:
            # Skip if already holding
            if any(p['symbol'] == stock['symbol'] for p in self.current_positions):
                continue
            
            # Check re-entry cooldown
            if not self.check_reentry_allowed(stock['symbol']):
                continue
            
            current_price = self.simulate_price_movement(stock['symbol'])
            base_price = stock['base_price']
            gain_pct = ((current_price - base_price) / base_price) * 100
            
            # Check if within adaptive entry range
            if quality['entry_min'] <= gain_pct <= quality['entry_max']:
                print(f"   📊 {stock['symbol']}: ₹{current_price} ({gain_pct:+.2f}%)")
                
                # Adaptive probability based on quality threshold
                if random.random() < quality['probability']:
                    success = self.execute_new_trade(stock)
                    if success:
                        return stock['symbol']
                else:
                    print(f"      ⏸️  Valid setup but waiting for higher conviction")
        
        return None
    
    def run_cycle(self, cycle_num):
        scan_time = self.get_current_time()
        
        print(f"\n{'='*70}")
        print(f"CYCLE #{cycle_num} - {scan_time} IST")
        print(f"{'='*70}")
        
        positions = self.load_positions()
        
        # Monitor existing positions
        for pos in positions:
            current_price = self.simulate_price_movement(pos['symbol'], pos['entry'])
            status, price = self.check_position_status(pos, current_price)
            
            pnl_unrealized = (current_price - pos['entry']) * pos['qty']
            pnl_pct = ((current_price - pos['entry']) / pos['entry']) * 100
            
            print(f"\n📊 {pos['symbol']}: ₹{current_price} (Entry: ₹{pos['entry']})")
            print(f"   Status: {status} | P&L: ₹{pnl_unrealized:+.0f} ({pnl_pct:+.2f}%)")
            
            if status in ["SL_HIT", "TARGET_HIT"]:
                pnl = self.update_position(pos, status, price)
                result = "✅ PROFIT" if pnl > 0 else "❌ LOSS"
                print(f"   {result}: ₹{pnl:+.2f} - Position CLOSED")
                self.current_positions.remove(pos)
            elif status == "TRAILING":
                print(f"   📈 In profit zone - trailing active")
        
        # Portfolio summary
        print(f"\n💼 Portfolio: {len(self.current_positions)}/{self.max_positions} positions")
        print(f"   Daily P&L: ₹{self.daily_pnl:+,.0f} | Total Trades: {self.trades_today}")
        
        # Check daily limits
        if self.daily_pnl <= -self.max_daily_loss:
            print("\n⛔ DAILY LOSS LIMIT HIT - STOP TRADING")
            return False
        
        if self.daily_pnl >= 5000:
            print("\n🎉 DAILY TARGET EXCEEDED (+5%) - STOP TRADING")
            return False
        
        # Scan for new entries
        new_trade = self.scan_and_execute()
        if new_trade:
            self.current_positions = self.load_positions()
        
        return True

def main():
    trader = AdaptiveTrader()
    
    print(f"""
{'='*70}
🤖 AUTONOMOUS TRADING SYSTEM v3.0 - ADAPTIVE QUALITY
{'='*70}
Monitor Interval: 10 minutes
No Daily Trade Limit (Quality increases with count)
Adaptive Entry Criteria: Trades 1-3 (wide) → 10+ (very tight)
Daily Loss Limit: -1.5% (₹1,500)
Defensive Sizing: Active after 2 consecutive losses
{'='*70}
    """)
    
    cycle = 1
    
    while True:
        try:
            continue_trading = trader.run_cycle(cycle)
            
            if not continue_trading:
                print("\n🛑 Trading stopped - limits reached")
                break
            
            current_hour = datetime.now().hour
            current_min = datetime.now().minute
            
            if current_hour >= 15 and current_min >= 15:
                print("\n🔔 Market closing - stopping monitor")
                break
            
            print(f"\n⏳ Next scan in 10 minutes...")
            # Sleep with heartbeat every minute
            for i in range(10):
                time.sleep(60)  # 1 minute
                mins_remaining = 10 - (i + 1)
                if mins_remaining > 0:
                    print(f"   ⏱️  {datetime.now().strftime('%H:%M:%S')} - {mins_remaining} min until next scan")
            
            cycle += 1
            
        except KeyboardInterrupt:
            print("\n\n⚠️ Monitor stopped by user")
            break
        except Exception as e:
            print(f"\n❌ Error in cycle: {e}")
            time.sleep(600)  # 10 minutes before retry
            continue
    
    print(f"""
{'='*70}
📊 TRADING SESSION COMPLETE
{'='*70}
Total Cycles: {cycle}
Final P&L: ₹{trader.daily_pnl:,.2f}
Total Trades: {trader.trades_today}
{'='*70}
    """)

if __name__ == "__main__":
    main()
