#!/usr/bin/env python3
"""Quick dashboard to check system status"""

from openpyxl import load_workbook
from datetime import datetime
import os

def print_dashboard():
    wb = load_workbook('trades.xlsx')
    sheet = wb['Trade Log']
    
    wins = losses = open_pos = 0
    total_pnl = 0
    
    for row in range(2, sheet.max_row + 1):
        outcome = sheet[f'O{row}'].value
        pnl = sheet[f'M{row}'].value or 0
        
        if outcome == "WIN":
            wins += 1
            total_pnl += pnl
        elif outcome == "LOSS":
            losses += 1
            total_pnl += pnl
        elif outcome == "OPEN":
            open_pos += 1
    
    status_emoji = "🟢" if total_pnl > 0 else "🔴" if total_pnl < -500 else "🟡"
    
    print(f"""
╔═══════════════════════════════════════════════════════════════════╗
║           🤖 AUTONOMOUS TRADING SYSTEM - DASHBOARD               ║
╚═══════════════════════════════════════════════════════════════════╝

  Time: {datetime.now().strftime('%H:%M:%S IST')}
  
  {status_emoji} Realized P&L: ₹{total_pnl:+,.2f}
  
  📊 Trades: {wins+losses} closed, {open_pos} open
  ✅ Wins: {wins}  |  ❌ Losses: {losses}
  📈 Win Rate: {(wins/(wins+losses)*100) if (wins+losses) > 0 else 0:.0f}%
  
  🎯 Trade Limit: {wins+losses+open_pos}/6
  🔒 Position Limit: {open_pos}/2

╔═══════════════════════════════════════════════════════════════════╗
║ OPEN POSITIONS                                                    ║
╚═══════════════════════════════════════════════════════════════════╝
""")
    
    if open_pos == 0:
        print("  No open positions - monitoring for entries...\n")
    else:
        for row in range(2, sheet.max_row + 1):
            if sheet[f'O{row}'].value == "OPEN":
                symbol = sheet[f'C{row}'].value
                entry = sheet[f'F{row}'].value
                sl = sheet[f'G{row}'].value
                target = sheet[f'H{row}'].value
                qty = sheet[f'I{row}'].value
                
                print(f"  📌 {symbol}: {qty} shares @ ₹{entry}")
                print(f"     SL: ₹{sl} | Target: ₹{target}\n")
    
    print(f"""╔═══════════════════════════════════════════════════════════════════╗
║ SYSTEM STATUS                                                     ║
╚═══════════════════════════════════════════════════════════════════╝

  Monitor: Running every 10 minutes
  Risk Limit: {abs(total_pnl)/1500*100:.1f}% of -₹1,500 daily limit
  Next Action: {'Continue monitoring' if open_pos > 0 else 'Scan for entries'}

═══════════════════════════════════════════════════════════════════════
""")

if __name__ == "__main__":
    print_dashboard()
